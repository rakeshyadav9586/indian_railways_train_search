from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
import selenium.webdriver.support.ui as ui
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from openpyxl import Workbook
from datetime import date
import datetime
import time
import os


excel_path = "../indian_railways_train_search/Excels/"
station_list = ["ADI"]
intersect_file = station_list[0] + "_" + station_list[1] + "_intersect.xlsx"
final_intersect_file = (
    "Final_" + station_list[0] + "_" + station_list[1] + "_intersect.xlsx"
)
wb = Workbook()
wb1 = Workbook()
start_time = time.time()
options = webdriver.ChromeOptions()
options.add_argument("--incognito")
# options.add_argument("--headless")
driver = webdriver.Chrome(options=options)
driver.implicitly_wait(10)

def station_wise_trains():
    global station_list, excel_path, wb, driver
    actionChains = ActionChains(driver)

    for station in range(0, len(station_list)):
        all_files = os.listdir(excel_path)
        if "stationwise_trains.xlsx" in all_files:
            wb = load_workbook(excel_path + "stationwise_trains.xlsx")
            print("excel file already there")
        else:
            wb.save(excel_path + "stationwise_trains.xlsx")
            print("new created and then loaded")
            wb = load_workbook(excel_path + "stationwise_trains.xlsx")

        if station_list[station] in wb.sheetnames:
            sheet = wb.active
        else:
            sheet = wb.create_sheet(station_list[station], 0)

        train_url = "https://etrain.info/in?STATION=" + station_list[station]
        print(train_url)
        driver.get(train_url)
        driver.maximize_window()
        wait = ui.WebDriverWait(driver, 10)
        wait.until(
            lambda drivers: driver.find_elements_by_xpath(
                '//div[@class="trainlist rnd5"]'
            )
        )
        total_train = driver.find_elements_by_xpath(
            '//div[@class="trainlist rnd5"]//tr'
        )
        for number in range(1, len(total_train) + 1):
            print(
                driver.find_element_by_xpath(
                    '//div[@class="trainlist rnd5"]//tr[' + str(number) + "]"
                ).text
            )
            row = driver.find_element_by_xpath(
                '//div[@class="trainlist rnd5"]//tr[' + str(number) + "]/td[1]"
            ).text
            sheet["A" + str(number)].value = row
        wb.save(excel_path + "stationwise_trains.xlsx")

        wb = load_workbook(excel_path + "stationwise_trains.xlsx")
        sheet = wb[station_list[station]]
        row_count = sheet.max_row
        print("Total rows : " + str(row_count))


def excel_create_train_sheet():
    global station_list, excel_path, wb, wb1, driver
    actionChains = ActionChains(driver)
    print(station_list)
    print("station list above")

    for station in range(0, len(station_list)):
        all_files = os.listdir(excel_path)
        if "stationwise_trains.xlsx" in all_files:
            wb = load_workbook(excel_path + "stationwise_trains.xlsx")
            print("excel file already there")
        else:
            wb.save(excel_path + "stationwise_trains.xlsx")
            print("new created and then loaded")
            wb = load_workbook(excel_path + "stationwise_trains.xlsx")

        if station_list[station] in wb.sheetnames:
            sheet = wb.active
        else:
            sheet = wb.create_sheet(station_list[station], 0)

        sheet = wb[station_list[station]]
        row_count = sheet.max_row
        print("Total rows : " + str(row_count))
        print(station_list[station])
        print("below station")
        excel_train_list = []

        for number in range(1, row_count + 1):
            cell = sheet["A" + str(number)]
            print(cell.value)
            excel_train_list.append(cell.value)
        print(excel_train_list)

        train_url = "https://etrain.info/in?TRAIN=" + excel_train_list[station]
        print(train_url)
        driver.get(train_url)
        # driver.maximize_window()
        total_station_row = driver.find_elements_by_xpath('//table[@id="schtbl"]//tr')
        print(len(total_station_row))

        stoppage_list = []
        all_files = os.listdir(excel_path)
        print(all_files)
        print(station_list[station])
        if station_list[station] + ".xlsx" in all_files:
            wb1 = load_workbook(excel_path + station_list[station] + ".xlsx")
            print("excel file already there")
        else:
            wb1.save(excel_path + station_list[station] + ".xlsx")
            print("new excel file created and then loaded")
            wb1 = load_workbook(excel_path + station_list[station] + ".xlsx")

        for row in range(3, len(total_station_row)):
            station_name = driver.find_element_by_xpath(
                '//table[@id="schtbl"]//tr[' + str(row) + "]/td[3]"
            ).text
            print(station_name)

    # need to update below code
    # if u"\u21D2" in header.text:
    #     raw = header.text
    #     first_symbol = raw.index(u"\u21D2")
    #     second_symbol = raw.index("/")
    #     update_train = str(raw[first_symbol + 1:second_symbol])
    #     print("if loop")
    #     print(update_train)
    # else:
    #     update_train = str(header.text[0:5])
    #     print("else loop")
    #     print(update_train)
    # if update_train in wb1.sheetnames:
    #     sheet1 = wb1.active
    #     print(update_train)
    #     print("sheet name already there")
    # else:
    #     sheet1 = wb1.create_sheet(update_train, 0)
    #     print("new sheet name created")
    # stoppage = driver.find_elements_by_xpath('//div[@class="newschtable newbg inline"]/div')
    # for no in range(3, len(stoppage) + 1, 3):
    #     stop = driver.find_element_by_xpath('//div[@class="newschtable newbg inline"]/div[' + str(no) + ']/div[4]').text
    #     # print stop
    #     stoppage_list.append(str(stop))
    # print(stoppage_list)
    # print("\n")
    # for row in range(0, len(stoppage_list)):
    #     sheet1["A" + str(row+1)].value = stoppage_list[row]
    # wb1.save(excel_path + station_list[station]+".xlsx")
    # stoppage_list = []

    time.sleep(2)


def match_station():

    global station_list, intersect_file, excel_path, wb
    print(station_list[0])
    print(station_list[1])
    wb_1 = load_workbook(excel_path + station_list[0] + ".xlsx")
    wb_2 = load_workbook(excel_path + station_list[1] + ".xlsx")

    print(wb_1.sheetnames)
    print(len(wb_1.sheetnames))
    for i in range(0, len(wb_1.sheetnames) - 1):
        print(wb_1.worksheets[i])
        sheet_1 = wb_1.worksheets[i]
        row_count_1 = sheet_1.max_row
        print(row_count_1)
        train_list_1 = []

        for number in range(1, row_count_1 + 1):
            cell_1 = sheet_1["A" + str(number)]
            print(cell_1.value)
            train_list_1.append(cell_1.value)
        print(train_list_1)

        print(wb_2.sheetnames)
        print(len(wb_2.sheetnames))
        for j in range(0, len(wb_2.sheetnames) - 1):
            train_list_2 = []
            print(wb_2.sheetnames[j])
            sheet_2 = wb_2.worksheets[j]
            row_count_2 = sheet_2.max_row
            print(row_count_2)

            for number in range(1, row_count_2 + 1):
                cell_2 = sheet_2["A" + str(number)]
                print(cell_2.value)
                train_list_2.append(cell_2.value)
            print(train_list_2)

            row_data = []
            matched = []
            raw_list = [station_list[0], sheet_1.title, station_list[1], sheet_2.title]
            for lis in raw_list:
                row_data.append(lis)
            for k in train_list_1:
                for l in train_list_2:
                    if k == l:
                        print(l)
                        matched.append(l)
                        count = train_list_2.index(l)
                        train_list_2 = train_list_2[count:]
            if matched:
                print("in matched if loop")
                print(matched)
                print(sheet_1.title)
                print(sheet_2.title)
                all_files = os.listdir(excel_path)
                if intersect_file in all_files:
                    wb = load_workbook(excel_path + intersect_file)
                    print("already there")
                else:
                    wb.save(excel_path + intersect_file)
                    print("new excel file created and then loaded")
                    wb = load_workbook(excel_path + intersect_file)
                sheet = wb.active
                sheet.append(row_data + matched)
                wb.save(excel_path + intersect_file)


def before_final_sheet():
    global excel_path, station_list, intersect_file, final_intersect_file, driver
    driver.maximize_window()
    actionChains = ActionChains(driver)
    gap_time = int(6)
    wb3 = Workbook()
    wb3 = load_workbook(excel_path + intersect_file)
    sheet3 = wb3.active
    row_count = sheet3.max_row
    intersect_train = []
    max_col = 5
    for j in range(1, row_count + 1):
        for i in range(1, max_col + 1):
            cell_obj = sheet3.cell(row=j, column=i)
            print("before split")
            print(cell_obj.value)
            row_value = str(cell_obj.value).split(" ")
            print(row_value)
            print("after split")
            print(row_value[0])
            intersect_train.append(row_value[0])
        print("for loop of row")
        print(intersect_train)
        print(intersect_train[1])
        print(intersect_train[3])
        intersect_station = str(intersect_train[4]).upper()
        trains = [intersect_train[1], intersect_train[3]]
        print(trains)
        default_time = 0
        print("after default_time")
        for t in trains:
            print(t)
            train_url = "https://etrain.info/in"
            print(train_url)
            driver.get(train_url)
            train_text = driver.find_element_by_xpath('//input[@id="bartrainid"]')
            train_text.clear()
            train_text.send_keys(t)
            time.sleep(1)
            driver.find_element_by_xpath('//a[@id="trnrtelnk"]').click()
            time.sleep(2)
            try:
                print("in try loop")
                table = driver.find_elements_by_xpath('//table[@id="schtbl"]//tr')
                print("checked train table")
                for i in range(3, len(table)):
                    row_data = driver.find_element_by_xpath(
                        '//table[@id="schtbl"]//tr[' + str(i) + "]"
                    ).text
                    print(row_data)
                    if intersect_station in row_data:
                        print("if loop for time check")
                        row_time = driver.find_element_by_xpath(
                            '//table[@id="schtbl"]//tr[' + str(i) + "]/td[4]"
                        ).text
                        if row_time == "Source":
                            print("Source station")
                            row_time = driver.find_element_by_xpath(
                                '//table[@id="schtbl"]//tr[' + str(i) + "]/td[5]"
                            ).text
                        print(row_time)
                        print("row time check")
                        int_time = int(row_time[:2])
                        print(int_time)

                        if (default_time >= 12) and (int_time <= 12):
                            ar = 24 - default_time
                            new_time = ar + int_time
                            print("if loop")
                            print(new_time)

                        elif (default_time >= 12) and (int_time >= 12):
                            ar = 24 - default_time
                            br = 24 - int_time
                            new_time = ar - br
                            if new_time < 0:
                                new_time = new_time + 24
                            print("elif loop")
                            print(new_time)
                        else:
                            new_time = int_time - default_time
                            if (new_time < 0) or (new_time == 0):
                                new_time = 24 + new_time
                            print("Else loop")
                            print(new_time)
                        new_time = int_time-default_time
                        print(new_time)

                        if new_time == int_time:
                            print("if loop new_time == int_time")
                            default_time = int_time
                        else:
                            print("else loop of time")
                            new_time = int(new_time)
                            print(new_time)
                            print("going to check <6")
                            if (new_time < gap_time) and (new_time != 0):
                                print("time gap is less than 6 hour")
                                print(trains[0])
                                print(trains[1])
                                print(intersect_station)
                                print("final train list")
                                final_list = [trains[0], trains[1], intersect_station]
                                print(final_list)
                                print("jo break")
                                print("Save below train list in excel")

                                all_files = os.listdir(excel_path)
                                wb4 = Workbook()
                                if final_intersect_file in all_files:
                                    wb4 = load_workbook(
                                        excel_path + final_intersect_file
                                    )
                                    print("already there")
                                else:
                                    wb4.save(excel_path + final_intersect_file)
                                    print("new excel file created and then loaded")
                                    wb4 = load_workbook(
                                        excel_path + final_intersect_file
                                    )
                                print("going to active sheet")
                                sheet4 = wb4.active
                                print("going to append sheet")
                                sheet4.append(final_list)
                                print("going to save sheet")
                                wb4.save(excel_path + final_intersect_file)
                                print("sheet saved")
                                final_list = []
                                print("empty final_list")

                        break
                intersect_train = []
            except NoSuchElementException:
                print(" error found")
                pass
    time.sleep(5)

# Run below method as per requirement
# station_wise_trains()
# excel_create_train_sheet()
# match_station()
# before_final_sheet()

driver.close()
driver.quit()
print("--- %s seconds ---" % (time.time() - start_time))
